# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import json
with open("./conf/config.json") as f:
    config = json.load(f)

hz = openpyxl.load_workbook(config['hzb'])#汇总表

wh= openpyxl.load_workbook(config['whb'])#考勤维护表
hzb=hz.active #读当前sheet
whb=wh.active
whbsy=[]#维护表索引
kbb=[]#空白表
#对照表
dzb={"A":1,
"B":2,
"C":3,
"D":4,
"E":5,
"F":6,
"G":7,
"H":8,
"I":9,
"J":10,
"K":11,
"L":12,
"M":13,
"N":14,
"O":15,
"P":16,
"Q":17,
"R":18,
"S":19,
"T":20,
"U":21,
"V":22,
"W":23,
"X":24,
"Y":25,
"Z":26,
"AA":27,
"AB":28,
"AC":29,
"AD":30,
"AE":31,
"AF":32,
"AG":33,
"AH":34,
"AI":35,
"AJ":36,
"AK":37,
"AL":38,
"AM":39,
"AN":40,
"AO":41,
"AP":42,
"AQ":43,
"AR":44,
"AS":45,
"AT":46,
"AU":47,
"AV":48,
"AW":49,
"AX":50,
"AY":51,
"AZ":52,
}

#建立维护表索引
for rows in whb.iter_rows(min_row=config['qsh'], min_col=config['syl'],
                          max_col=config['syl']):
    whbsy.append( rows[0])

#统计空白点
for rows in hzb.iter_rows(min_row=config['qsh'], min_col=config['qsl'], max_col=config['dyts']+config['qsl']):
    for kongbai in rows:
        if kongbai.value==None or kongbai.value == '':
            kbb.append(kongbai)

#通过索引将维护表数据读取出来将考勤表中空白部分补全
for a in kbb:
    for sy in whbsy:
        if sy.value  ==hzb.cell(a.row,config['syl']).value:
            a.value = whb.cell(sy.row,dzb[a.column]).value
#存储excel
hz.save(config['new_name'])
